from openpyxl import load_workbook
import json

status_sheet = "./VCodes2016-17.xlsx"


# route has vcode has cip
def main(sheet=status_sheet):
    wb = load_workbook(sheet)
    ws = wb.active
    max_row = ws.max_row
    routes = []
    route = None
    for row in range(1, max_row):
        if ws.cell(row, 1).value:
            # next route?
            if not ws.cell(row, 2).value:
                if not route:
                    route = {"name": ws.cell(row, 1).value}
                else:
                    route["type"] = ws.cell(row, 1).value
                    route["vcodes"] = []
                    routes.append(route)
                    route = None
            else:
                if ws.cell(row, 1).value == "CERTIFICATION CODE":
                    continue
                if ws.cell(row, 1).value:
                    vcode = {
                        "vcode": ws.cell(row, 1).value.strip(),
                        "certification": ws.cell(row, 2).value.strip(),
                        "experience": ws.cell(row, 3).value.strip(),
                        "cips": [],
                    }
                    routes[-1]["vcodes"].append(vcode)
        if ws.cell(row, 4).value:
            cip = {
                "code": ws.cell(row, 4).value,
                "name": ws.cell(row, 5).value.strip(),
            }
            vcode["cips"].append(cip)

    # filter to plan 2 routes
    business = [route for route in routes if "Industry" in route["type"]]

    # transform from by-vcode to by-cip
    bycip = dict()
    for route in business:
        for vcode in route["vcodes"]:
            for cip in vcode["cips"]:
                cipObj = bycip.get(
                    cip["code"], {"code": cip["code"], "name": cip["name"], "vcode": []}
                )
                cipObj["vcode"].append(
                    {"vcode": vcode["vcode"], "certification": vcode["certification"]}
                )
                bycip[cipObj["code"]] = cipObj

    # filter cips with multiple vcodes
    multis = [cip for cip in bycip.values() if len(cip["vcode"]) > 1]

    with open("vcode.json", "w") as fp:
        json.dump(multis, fp)


# then run: python -m http.server
# and open http://localhost:8000/vcode.html

if __name__ == "__main__":
    main()
